from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

from pmao.models import Initiative

# ── Column definitions ─────────────────────────────────────────────────────────

INITIATIVE_COLS = [
    "initiative_id",
    "initiative",
    "coordination_owner",
    "responsible_owner",
    "status",
    "priority",
    "current_state",
    "coordination_next_steps",
    "outstanding_questions",
    "outstanding_meetings",
    "last_touch_comment",
    "last_touch_timestamp",
    "syndication_notes",
    "materials_link",
    "notes",
]

NEXT_STEPS_LOG_COLS = ["initiative_id", "initiative", "coordination_next_steps", "date_captured"]

LAST_TOUCH_LOG_COLS = ["initiative_id", "initiative", "last_touch_timestamp", "last_touch_comment"]

MEETING_REQUEST_COLS = [
    "initiative_ids", "initiative", "meeting_title", "duration",
    "suggested_date", "attendees", "scheduler_responsible", "notes", "status",
]

MILESTONE_COLS = [
    "date", "end_date", "title", "type", "priority",
    "initiative_ids", "context", "status",
]

HYPOTHESIS_COLS = [
    "id", "level", "initiative", "hypothesis", "owner",
    "validation_path", "status", "created", "source",
]

FACT_COLS = [
    "id", "initiative", "claim", "stated_by", "authority",
    "confidence", "inferred", "source_span", "source", "created",
]

SIGNAL_COLS = [
    "id", "initiative", "principal", "lever", "signal",
    "implication", "source", "created",
]

MEETING_COLS = [
    "id", "initiative", "purpose", "convener", "attendees",
    "functions", "target_timing", "cadence", "status", "row_source", "created",
]

REVIEW_QUEUE_COLS = ["staging_file", "category", "initiative_id", "summary"]

# ── Styles ─────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_GRAY   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_THIN   = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _write_header_row(ws, cols: list) -> None:
    for ci, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[1].height = 30


def _write_data_row(ws, row_num: int, cols: list, record: dict) -> None:
    for ci, col_name in enumerate(cols, start=1):
        val = record.get(col_name, "") or ""
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = _BORDER


def _add_status_cf(ws, status_col_idx: int, max_row: int, num_cols: int) -> None:
    status_letter = get_column_letter(status_col_idx)
    data_range = f"A2:{get_column_letter(num_cols)}{max_row}"
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[f'${status_letter}2="ready"'], fill=_GREEN,
    ))
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[f'${status_letter}2="in_progress"'], fill=_YELLOW,
    ))
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[f'${status_letter}2="complete"'], fill=_GRAY,
    ))


def create_workbook(
    path: Path,
    initiatives: List[Initiative],
    meeting_requests: Optional[list] = None,
    milestones: Optional[list] = None,
    hypotheses: Optional[list] = None,
    facts: Optional[list] = None,
    signals: Optional[list] = None,
    meetings: Optional[list] = None,
    staged: Optional[list] = None,
) -> None:
    wb = Workbook()

    # ── Tab 1: Initiatives ────────────────────────────────────────────────────
    ws_init = wb.active
    ws_init.title = "Initiatives"
    _write_header_row(ws_init, INITIATIVE_COLS)

    for row_num, init in enumerate(initiatives, start=2):
        record = {
            "initiative_id": init.id,
            "initiative": init.name,
            "coordination_owner": init.coordination_owner or "",
            "responsible_owner": init.responsible_owner or "",
            "status": init.status,
            "priority": init.priority or "",
            "current_state": init.current_state or "",
            "coordination_next_steps": init.coordination_next_steps or "",
            "outstanding_questions": init.outstanding_questions or "",
            "outstanding_meetings": init.outstanding_meetings or "",
            "last_touch_comment": init.last_touch_comment or "",
            "last_touch_timestamp": init.last_touch_timestamp.isoformat() if init.last_touch_timestamp else "",
            "syndication_notes": init.syndication_notes or "",
            "materials_link": init.materials_link or "",
            "notes": init.notes or "",
        }
        _write_data_row(ws_init, row_num, INITIATIVE_COLS, record)

    status_col = INITIATIVE_COLS.index("status") + 1
    _add_status_cf(ws_init, status_col, max(2, 1 + len(initiatives)), len(INITIATIVE_COLS))

    ws_init.column_dimensions["A"].width = 10
    ws_init.column_dimensions["B"].width = 32
    ws_init.column_dimensions["C"].width = 24
    ws_init.column_dimensions["D"].width = 24
    ws_init.column_dimensions["E"].width = 14
    ws_init.column_dimensions["F"].width = 10
    ws_init.column_dimensions["G"].width = 40
    ws_init.column_dimensions["H"].width = 40
    ws_init.column_dimensions["I"].width = 38
    ws_init.column_dimensions["J"].width = 38
    ws_init.column_dimensions["K"].width = 40
    ws_init.column_dimensions["L"].width = 16
    ws_init.column_dimensions["M"].width = 30
    ws_init.column_dimensions["N"].width = 30
    ws_init.column_dimensions["O"].width = 40
    ws_init.freeze_panes = "B2"

    # ── Tab 2: Full Next Steps Log ────────────────────────────────────────────
    ws_log = wb.create_sheet("Full Next Steps Log", index=1)
    _write_header_row(ws_log, NEXT_STEPS_LOG_COLS)
    for row_num, init in enumerate(initiatives, start=2):
        _write_data_row(ws_log, row_num, NEXT_STEPS_LOG_COLS, {
            "initiative_id": init.id,
            "initiative": init.name,
            "coordination_next_steps": init.coordination_next_steps or "",
            "date_captured": init.last_touched.isoformat(),
        })
    ws_log.column_dimensions["A"].width = 12
    ws_log.column_dimensions["B"].width = 32
    ws_log.column_dimensions["C"].width = 80
    ws_log.column_dimensions["D"].width = 14
    ws_log.freeze_panes = "A2"

    # ── Tab 3: Last Touch Log ─────────────────────────────────────────────────
    ws_lt = wb.create_sheet("Last Touch Log", index=2)
    _write_header_row(ws_lt, LAST_TOUCH_LOG_COLS)
    for row_num, init in enumerate(initiatives, start=2):
        _write_data_row(ws_lt, row_num, LAST_TOUCH_LOG_COLS, {
            "initiative_id": init.id,
            "initiative": init.name,
            "last_touch_timestamp": init.last_touch_timestamp.isoformat() if init.last_touch_timestamp else "",
            "last_touch_comment": init.last_touch_comment or "",
        })
    ws_lt.column_dimensions["A"].width = 12
    ws_lt.column_dimensions["B"].width = 32
    ws_lt.column_dimensions["C"].width = 16
    ws_lt.column_dimensions["D"].width = 80
    ws_lt.freeze_panes = "A2"

    # ── Tab 4: Meeting Requests ───────────────────────────────────────────────
    ws_mr = wb.create_sheet("Meeting Requests")
    _write_header_row(ws_mr, MEETING_REQUEST_COLS)
    if meeting_requests:
        for row_num, mr in enumerate(meeting_requests, start=2):
            _write_data_row(ws_mr, row_num, MEETING_REQUEST_COLS, mr)
    ws_mr.freeze_panes = "A2"

    # ── Tab 5: Milestones ─────────────────────────────────────────────────────
    ws_ms = wb.create_sheet("Milestones")
    _write_header_row(ws_ms, MILESTONE_COLS)
    if milestones:
        for row_num, ms in enumerate(milestones, start=2):
            _write_data_row(ws_ms, row_num, MILESTONE_COLS, ms)
    ws_ms.freeze_panes = "A2"

    init_names = {i.id: i.name for i in initiatives}

    # ── Tab 6: Hypotheses ─────────────────────────────────────────────────────
    ws_hyp = wb.create_sheet("Hypotheses")
    _write_header_row(ws_hyp, HYPOTHESIS_COLS)

    if hypotheses:
        # Sort: project-level (initiative_id is None) first, then initiative-level by name
        project_rows = [h for h in hypotheses if not h.get("initiative_id")]
        initiative_rows = sorted(
            [h for h in hypotheses if h.get("initiative_id")],
            key=lambda h: init_names.get(h["initiative_id"], h["initiative_id"]),
        )
        sorted_hyp = project_rows + initiative_rows

        for row_num, h in enumerate(sorted_hyp, start=2):
            iid = h.get("initiative_id")
            record = {
                "id": h.get("id", ""),
                "level": "Project" if not iid else "Initiative",
                "initiative": init_names.get(iid, "") if iid else "",
                "hypothesis": h.get("hypothesis", ""),
                "owner": h.get("owner", ""),
                "validation_path": h.get("validation_path", ""),
                "status": h.get("status", ""),
                "created": h.get("created", ""),
                "source": h.get("source", ""),
            }
            _write_data_row(ws_hyp, row_num, HYPOTHESIS_COLS, record)

    ws_hyp.column_dimensions["A"].width = 22
    ws_hyp.column_dimensions["B"].width = 12
    ws_hyp.column_dimensions["C"].width = 32
    ws_hyp.column_dimensions["D"].width = 60
    ws_hyp.column_dimensions["E"].width = 20
    ws_hyp.column_dimensions["F"].width = 40
    ws_hyp.column_dimensions["G"].width = 14
    ws_hyp.column_dimensions["H"].width = 12
    ws_hyp.column_dimensions["I"].width = 30
    ws_hyp.freeze_panes = "A2"

    # ── Tab 7: Facts ──────────────────────────────────────────────────────────
    ws_f = wb.create_sheet("Facts")
    _write_header_row(ws_f, FACT_COLS)
    for row_num, f in enumerate(facts or [], start=2):
        iid = f.get("initiative_id")
        record = dict(f)
        record["initiative"] = init_names.get(iid, iid or "")
        record["inferred"] = "yes" if f.get("inferred") else "no"
        _write_data_row(ws_f, row_num, FACT_COLS, record)
    ws_f.column_dimensions["B"].width = 28
    ws_f.column_dimensions["C"].width = 60
    ws_f.column_dimensions["H"].width = 40
    ws_f.freeze_panes = "A2"

    # ── Tab 8: Signals ────────────────────────────────────────────────────────
    ws_s = wb.create_sheet("Signals")
    _write_header_row(ws_s, SIGNAL_COLS)
    for row_num, s in enumerate(signals or [], start=2):
        iid = s.get("initiative_id")
        record = dict(s)
        record["initiative"] = init_names.get(iid, iid or "")
        _write_data_row(ws_s, row_num, SIGNAL_COLS, record)
    ws_s.column_dimensions["B"].width = 28
    ws_s.column_dimensions["E"].width = 50
    ws_s.column_dimensions["F"].width = 50
    ws_s.freeze_panes = "A2"

    # ── Tab 9: Meetings (structured + initiative free-text, combined) ────────
    ws_m = wb.create_sheet("Meetings")
    _write_header_row(ws_m, MEETING_COLS)
    meeting_rows = []
    for m in meetings or []:
        iid = m.get("initiative_id")
        record = dict(m)
        record["initiative"] = init_names.get(iid, iid or "")
        record["attendees"] = ", ".join(m.get("attendees") or [])
        record["functions"] = ", ".join(m.get("functions") or [])
        record["row_source"] = "meetings.json"
        meeting_rows.append(record)
    for init in initiatives:
        if init.outstanding_meetings:
            meeting_rows.append({
                "id": "", "initiative": init.name,
                "purpose": init.outstanding_meetings,
                "convener": "", "attendees": "", "functions": "",
                "target_timing": "", "cadence": "", "status": "open",
                "row_source": "initiative-field", "created": "",
            })
    for row_num, record in enumerate(meeting_rows, start=2):
        _write_data_row(ws_m, row_num, MEETING_COLS, record)
    ws_m.column_dimensions["B"].width = 28
    ws_m.column_dimensions["C"].width = 60
    ws_m.column_dimensions["E"].width = 30
    ws_m.freeze_panes = "A2"

    # ── Tab 10: Review Queue (pending staged items) ───────────────────────────
    ws_rq = wb.create_sheet("Review Queue")
    _write_header_row(ws_rq, REVIEW_QUEUE_COLS)
    for row_num, item in enumerate(staged or [], start=2):
        _write_data_row(ws_rq, row_num, REVIEW_QUEUE_COLS, item)
    ws_rq.column_dimensions["A"].width = 30
    ws_rq.column_dimensions["B"].width = 20
    ws_rq.column_dimensions["C"].width = 14
    ws_rq.column_dimensions["D"].width = 80
    ws_rq.freeze_panes = "A2"

    wb.save(path)
