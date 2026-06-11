import json
import tempfile
import yaml
from datetime import date
from pathlib import Path

from pmao.models import Initiative
from pmao.vault import init_vault, load_initiatives, save_initiatives, VAULT_FILES


def _make_initiative(id_="init-001", name="Customer Data Platform"):
    return Initiative(
        id=id_, name=name, status="not_started",
        created=date(2026, 6, 8), last_touched=date(2026, 6, 8)
    )


def test_init_vault_creates_expected_files():
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        for fname in VAULT_FILES:
            assert (vault / fname).exists(), f"Missing {fname}"
        assert (vault / "workbook.xlsx").exists()
        assert (vault / "transcripts").is_dir()
        assert (vault / "hypotheses.json").exists()


def test_init_vault_writes_config_with_project_name():
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault, project_name="Acme LRP", owner="J. Smith")
        cfg = yaml.safe_load((vault / "project-config.yaml").read_text())
        assert cfg["project_name"] == "Acme LRP"
        assert cfg["owner"] == "J. Smith"


def test_init_vault_creates_empty_initiatives():
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        data = json.loads((vault / "initiatives.json").read_text())
        assert data == []


def test_save_and_load_round_trip():
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        initiative = _make_initiative()
        save_initiatives(vault, [initiative])
        loaded = load_initiatives(vault)
        assert len(loaded) == 1
        assert loaded[0].id == "init-001"
        assert loaded[0].name == "Customer Data Platform"


def test_save_creates_backup():
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        save_initiatives(vault, [_make_initiative()])
        save_initiatives(vault, [_make_initiative("init-002", "Second")])
        assert (vault / "initiatives.json.bak").exists()


def test_no_lrp_in_config():
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        config_text = (vault / "project-config.yaml").read_text()
        assert "lrp" not in config_text.lower()
        assert "vault_type" not in config_text


def test_init_vault_creates_hypotheses_json():
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        hyp_path = vault / "hypotheses.json"
        assert hyp_path.exists()
        assert json.loads(hyp_path.read_text()) == []


def test_init_vault_creates_deep_extraction_files():
    import tempfile
    from pathlib import Path
    from pmao.vault import init_vault
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        assert (vault / "staging").is_dir()
        assert (vault / "facts.json").read_text() == "[]"
        assert (vault / "signals.json").read_text() == "[]"
        assert (vault / "meetings.json").read_text() == "[]"


def test_load_list_missing_file_returns_empty():
    import tempfile
    from pathlib import Path
    from pmao.vault import load_list
    with tempfile.TemporaryDirectory() as tmp:
        assert load_list(Path(tmp), "facts.json") == []


def test_load_list_reads_json_array():
    import json
    import tempfile
    from pathlib import Path
    from pmao.vault import load_list
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        (vault / "facts.json").write_text(json.dumps([{"id": "fact-1"}]))
        assert load_list(vault, "facts.json") == [{"id": "fact-1"}]


def test_staging_summaries_flattens_pending_files():
    import json
    import tempfile
    from pathlib import Path
    from pmao.vault import init_vault, staging_summaries
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        staged = {
            "status": "pending_review",
            "ingested": "2026-06-10",
            "source": "sync.vtt",
            "prompt_version": "pmao-deep-v1.0",
            "extraction": {
                "facts": [{"initiative_id": "init-001", "claim": "cost up 9%"}],
                "action_items": [{"initiative_id": None, "description": "build model"}],
                "review_flags": ["possible board-memo commitment, owner unclear"],
            },
        }
        (vault / "staging" / "2026-06-10-sync.json").write_text(json.dumps(staged))
        # Reviewed files are excluded
        done = dict(staged, status="reviewed")
        (vault / "staging" / "2026-06-09-old.json").write_text(json.dumps(done))

        rows = staging_summaries(vault)
        assert len(rows) == 3  # fact + action + review_flag
        by_cat = {r["category"]: r for r in rows}
        assert by_cat["facts"]["summary"] == "cost up 9%"
        assert by_cat["facts"]["staging_file"] == "2026-06-10-sync.json"
        assert by_cat["action_items"]["initiative_id"] == ""
        assert by_cat["review_flag"]["summary"].startswith("possible board-memo")


def test_staging_summaries_skips_corrupt_file():
    """Fix 2: a corrupt staging JSON is skipped with a warning; valid rows still returned."""
    import tempfile
    from pathlib import Path
    from pmao.vault import init_vault, staging_summaries, refresh_workbook
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        # corrupt file
        (vault / "staging" / "0000-corrupt.json").write_text("{not json")
        # valid pending file
        valid_payload = {
            "status": "pending_review",
            "ingested": "2026-06-10",
            "source": "good.txt",
            "prompt_version": "v",
            "extraction": {
                "facts": [{"initiative_id": "init-001", "claim": "GM is 62%",
                           "stated_by": "Sarah"}],
            },
        }
        (vault / "staging" / "2026-06-10-good.json").write_text(json.dumps(valid_payload))
        rows = staging_summaries(vault)
        assert len(rows) == 1
        assert rows[0]["summary"] == "GM is 62%"
        # refresh_workbook must not raise either
        refresh_workbook(vault)


def test_refresh_workbook_renders_canonical_and_staged():
    import json
    import tempfile
    from pathlib import Path
    from openpyxl import load_workbook
    from pmao.vault import init_vault, refresh_workbook
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        (vault / "facts.json").write_text(json.dumps(
            [{"id": "fact-0001", "initiative_id": "general", "claim": "a fact"}]))
        (vault / "staging" / "x.json").write_text(json.dumps({
            "status": "pending_review",
            "extraction": {"facts": [{"initiative_id": "general", "claim": "staged fact"}]},
        }))
        refresh_workbook(vault)
        wb = load_workbook(vault / "workbook.xlsx")
        assert wb["Facts"].cell(row=2, column=3).value == "a fact"
        assert wb["Review Queue"].cell(row=2, column=4).value == "staged fact"


def test_seed_from_csv_missing_name_column_raises_clear_error():
    import pytest
    from pmao.vault import seed_from_csv
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        csv_path = vault / "roster.csv"
        csv_path.write_text("id,coordination_owner\ninit-001,Alex Jordan\n")
        with pytest.raises(ValueError, match="name"):
            seed_from_csv(vault, csv_path)
