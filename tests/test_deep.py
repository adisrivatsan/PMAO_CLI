from pathlib import Path

SKILL_PATH = Path(__file__).parent.parent / "pmao" / "skills" / "ingest-deep.md"
PLACEHOLDERS = ["{initiatives}", "{roster}", "{reconciliation_candidates}", "{calibration}", "{source}"]


def test_skill_file_has_all_placeholders():
    text = SKILL_PATH.read_text(encoding="utf-8")
    for ph in PLACEHOLDERS:
        assert ph in text, f"missing placeholder {ph}"


def test_skill_file_has_no_legacy_roster_or_chunking():
    text = SKILL_PATH.read_text(encoding="utf-8")
    assert "owner_roster" not in text
    assert "principal_roster" not in text
    assert "pre-chunked" not in text
    assert '"chunk": "1/1"' in text
    assert "pmao-deep-v1.1" in text


def test_skill_file_requires_initiative_id_everywhere():
    text = SKILL_PATH.read_text(encoding="utf-8")
    # Every category block in the output schema carries initiative_id
    assert text.count('"initiative_id"') >= 7


import json
import tempfile
from datetime import date

from pmao.models import Initiative


def _make_initiative(id_="init-001", name="Pricing"):
    return Initiative(id=id_, name=name, status="in_progress",
                      created=date(2026, 6, 1), last_touched=date(2026, 6, 10))


def test_build_deep_prompt_fills_all_placeholders():
    from pmao.deep import _build_deep_prompt
    skill = SKILL_PATH.read_text(encoding="utf-8")
    prompt = _build_deep_prompt(
        skill, [_make_initiative()],
        roster_text="- Sarah Klein | owns: finance",
        recon_text="none",
        calibration_text="No prior calibration.",
        source_text="MEETING TEXT HERE",
    )
    for ph in PLACEHOLDERS:
        assert ph not in prompt, f"unfilled placeholder {ph}"
    assert "init-001: Pricing (status: in_progress)" in prompt
    assert "Sarah Klein" in prompt
    assert "MEETING TEXT HERE" in prompt


def test_reconciliation_block_filters_and_caps():
    from pathlib import Path
    from pmao.deep import _reconciliation_block, RECON_DECISION_CAP
    from pmao.vault import init_vault
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        actions = [
            {"id": "act-1", "description": "open one", "owner": "Dev Patel", "due": "", "status": "open"},
            {"id": "act-2", "description": "closed one", "owner": "Dev Patel", "due": "", "status": "done"},
        ]
        (vault / "actions.json").write_text(json.dumps(actions))
        decisions = [{"id": f"dec-{i}", "decision": f"decision {i}"} for i in range(30)]
        (vault / "decisions.json").write_text(json.dumps(decisions))

        block = _reconciliation_block(vault)
        assert "act-1: open one" in block
        assert "closed one" not in block
        assert "dec-29" in block          # newest kept
        assert "dec-0:" not in block      # oldest dropped by the cap
        assert block.count("dec-") == RECON_DECISION_CAP


def test_reconciliation_block_empty_says_none():
    from pathlib import Path
    from pmao.deep import _reconciliation_block
    from pmao.vault import init_vault
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        assert _reconciliation_block(vault) == "none"


def test_calibration_block_absent_and_capped():
    from pathlib import Path
    from pmao.deep import _calibration_block, CALIBRATION_LINE_CAP
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        assert _calibration_block(vault) == "No prior calibration."
        (vault / "learning").mkdir()
        lines = [f"lesson {i}" for i in range(300)]
        (vault / "learning" / "calibration.md").write_text("\n".join(lines))
        block = _calibration_block(vault)
        assert "lesson 299" in block      # newest kept
        assert "lesson 0\n" not in block  # oldest dropped
        assert len(block.splitlines()) == CALIBRATION_LINE_CAP


def test_validate_extraction_flags_unknown_initiative_and_keeps_item():
    from pmao.deep import _validate_extraction
    extraction = {
        "facts": [{"initiative_id": "init-999", "claim": "orphan claim"}],
        "hypotheses": [{"initiative_id": None, "theory": "project-level ok"}],
        "action_items": [{"initiative_id": "general", "description": "general ok"}],
    }
    out = _validate_extraction(extraction, [_make_initiative()])
    assert len(out["facts"]) == 1                      # kept, never dropped
    assert out["decisions"] == []                      # missing keys defaulted
    assert out["meetings_to_schedule"] == []
    flags = out["review_flags"]
    assert any("init-999" in f for f in flags)
    assert not any("general" in f for f in flags)
    assert len(flags) == 1                             # null and "general" not flagged


def test_validate_extraction_defaults_meta_and_preserves_existing():
    from pmao.deep import _validate_extraction
    out = _validate_extraction({}, [_make_initiative()])
    assert out["meta"] == {}                           # missing meta defaulted
    existing = {"meta": {"meeting_slug": "pricing-sync"}}
    out2 = _validate_extraction(existing, [_make_initiative()])
    assert out2["meta"] == {"meeting_slug": "pricing-sync"}  # not clobbered


def test_write_staging_shape_and_collision_suffix():
    from pathlib import Path
    from pmao.deep import _write_staging, PROMPT_VERSION
    from pmao.vault import init_vault
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        extraction = {"facts": [{"initiative_id": "general", "claim": "x"}]}
        p1 = _write_staging(vault, extraction, "Pricing Sync.vtt")
        p2 = _write_staging(vault, extraction, "Pricing Sync.vtt")
        assert p1.name == f"{date.today().isoformat()}-pricing-sync.json"
        assert p2.name == f"{date.today().isoformat()}-pricing-sync-2.json"
        staged = json.loads(p1.read_text())
        assert staged["status"] == "pending_review"
        assert staged["source"] == "Pricing Sync.vtt"
        assert staged["prompt_version"] == PROMPT_VERSION
        assert staged["extraction"] == extraction       # stored verbatim


EXTRACTION_FIXTURE = {
    "meta": {"meeting_slug": "2026-06-10-sync", "prompt_version": "pmao-deep-v1.1", "chunk": "1/1"},
    "facts": [{"initiative_id": "init-001", "claim": "cost up 9%", "stated_by": "Sarah Klein",
               "authority": "owner", "confidence": "high", "inferred": False, "source_span": "l12"}],
    "action_items": [{"initiative_id": "init-001", "type": "analysis_required",
                      "description": "build the cost model", "owner": "Dev Patel",
                      "owner_resolution": "alias_merged", "due": "unspecified", "source_span": "l40"}],
    "alias_flags": [{"variants": ["Dev"], "resolved_to": "Dev Patel", "confidence": "high", "needs_review": True}],
    "review_flags": ["near-discard: vague SLA aside"],
    "discard_note": "Discarded ~80%: logistics",
}


def _deep_vault(tmp):
    from pathlib import Path
    from pmao.vault import init_vault, save_initiatives
    vault = Path(tmp)
    init_vault(vault)
    save_initiatives(vault, [_make_initiative()])
    return vault


def test_run_ingest_deep_stages_without_touching_canonical(capsys):
    from unittest.mock import patch
    from openpyxl import load_workbook
    from pmao.deep import run_ingest_deep
    with tempfile.TemporaryDirectory() as tmp:
        vault = _deep_vault(tmp)
        source = vault / "transcripts" / "sync.txt"
        source.write_text("Sarah: cost up 9%. Dev, build the model.")

        with patch("pmao.llm.call_structured", return_value=dict(EXTRACTION_FIXTURE)) as mock:
            run_ingest_deep(vault, source)

        # Prompt was fully assembled (no leftover placeholders)
        prompt = mock.call_args[0][0]
        for ph in PLACEHOLDERS:
            assert ph not in prompt

        # Staged file exists with verbatim extraction
        staged_files = list((vault / "staging").glob("*.json"))
        assert len(staged_files) == 1
        staged = json.loads(staged_files[0].read_text())
        assert staged["status"] == "pending_review"
        assert staged["extraction"]["facts"][0]["claim"] == "cost up 9%"

        # Canonical files untouched
        assert json.loads((vault / "actions.json").read_text()) == []
        assert json.loads((vault / "facts.json").read_text()) == []

        # Workbook Review Queue reflects staging
        wb = load_workbook(vault / "workbook.xlsx")
        queue = [wb["Review Queue"].cell(row=r, column=4).value for r in (2, 3, 4)]
        assert "cost up 9%" in queue
        assert "build the cost model" in queue

        out = capsys.readouterr().out
        assert "staged, not applied" in out
        assert "near-discard: vague SLA aside" in out
        assert "pmao review" in out


def test_run_ingest_deep_empty_extraction_stages_nothing(capsys):
    from unittest.mock import patch
    from pmao.deep import run_ingest_deep
    with tempfile.TemporaryDirectory() as tmp:
        vault = _deep_vault(tmp)
        source = vault / "transcripts" / "sync.txt"
        source.write_text("nothing of substance")
        with patch("pmao.llm.call_structured", return_value={}):
            run_ingest_deep(vault, source)
        assert list((vault / "staging").glob("*.json")) == []
        assert "Nothing extracted" in capsys.readouterr().out


def test_run_ingest_deep_warns_when_roster_missing(capsys):
    from unittest.mock import patch
    from pmao.deep import run_ingest_deep
    with tempfile.TemporaryDirectory() as tmp:
        vault = _deep_vault(tmp)  # no roster.yaml written
        source = vault / "transcripts" / "sync.txt"
        source.write_text("text")
        with patch("pmao.llm.call_structured", return_value={}):
            run_ingest_deep(vault, source)
        assert "no roster.yaml" in capsys.readouterr().out


def test_cli_ingest_deep_flag_dispatches():
    import sys
    from unittest.mock import patch
    from pmao.cli import main
    with patch("pmao.deep.run_ingest_deep") as mock_deep, \
         patch.object(sys, "argv", ["pmao", "ingest", "v/", "--source", "s.txt", "--deep"]):
        main()
    mock_deep.assert_called_once()
    kwargs = mock_deep.call_args.kwargs
    assert str(kwargs["vault_path"]) == "v"
    assert str(kwargs["source_path"]) == "s.txt"


def test_run_ingest_deep_stages_flags_only_extraction(capsys):
    from unittest.mock import patch
    from pmao.deep import run_ingest_deep
    with tempfile.TemporaryDirectory() as tmp:
        vault = _deep_vault(tmp)
        source = vault / "transcripts" / "sync.txt"
        source.write_text("mostly noise")
        with patch("pmao.llm.call_structured",
                   return_value={"review_flags": ["near-discard: vague aside"]}):
            run_ingest_deep(vault, source)
        staged_files = list((vault / "staging").glob("*.json"))
        assert len(staged_files) == 1          # flags alone now stage
        assert "Staged for review" in capsys.readouterr().out


def test_run_ingest_deep_stages_alias_only_extraction(capsys):
    from unittest.mock import patch
    from pmao.deep import run_ingest_deep
    with tempfile.TemporaryDirectory() as tmp:
        vault = _deep_vault(tmp)
        source = vault / "transcripts" / "sync.txt"
        source.write_text("Dev said hi")
        alias_only = {"alias_flags": [{"variants": ["Dev"], "resolved_to": "Dev Patel",
                                       "confidence": "high", "needs_review": True}]}
        with patch("pmao.llm.call_structured", return_value=alias_only):
            run_ingest_deep(vault, source)
        staged_files = list((vault / "staging").glob("*.json"))
        assert len(staged_files) == 1          # aliases alone now stage
        staged = json.loads(staged_files[0].read_text())
        assert staged["extraction"]["alias_flags"][0]["resolved_to"] == "Dev Patel"
        out = capsys.readouterr().out
        assert "Staged for review" in out
        assert "Nothing extracted" not in out


def test_run_ingest_deep_sanitizes_control_chars_and_dict_values():
    from unittest.mock import patch
    from pmao.deep import run_ingest_deep
    from pmao.vault import refresh_workbook
    with tempfile.TemporaryDirectory() as tmp:
        vault = _deep_vault(tmp)
        source = vault / "transcripts" / "sync.txt"
        source.write_text("hostile content")
        hostile = {"facts": [
            {"initiative_id": "init-001", "claim": "budget is \x01 fine"},
            {"initiative_id": "init-001", "claim": {"text": "nested"}},
        ]}
        with patch("pmao.llm.call_structured", return_value=hostile):
            run_ingest_deep(vault, source)    # must not crash refresh_workbook
        staged_files = list((vault / "staging").glob("*.json"))
        assert len(staged_files) == 1
        staged = json.loads(staged_files[0].read_text())
        claims = [f["claim"] for f in staged["extraction"]["facts"]]
        assert claims[0] == "budget is  fine"            # control char stripped
        assert isinstance(claims[1], str)                # dict coerced to str
        assert "nested" in claims[1]
        refresh_workbook(vault)                          # export stays healthy


def test_validate_extraction_rejects_non_dict_top_level():
    import pytest
    from pmao.deep import _validate_extraction
    with pytest.raises(ValueError):
        _validate_extraction([], [_make_initiative()])


def test_validate_extraction_normalizes_wrong_shapes():
    from pmao.deep import _validate_extraction
    extraction = {
        "facts": None,
        "hypotheses": "not a list",
        "decisions": [None, "raw string", {"initiative_id": ["init-001"], "decision": "ok"}],
        "alias_flags": None,
        "review_flags": None,
        "meta": None,
    }
    out = _validate_extraction(extraction, [_make_initiative()])
    assert out["facts"] == []
    assert out["hypotheses"] == []
    assert out["decisions"] == [{"initiative_id": ["init-001"], "decision": "ok"}]
    assert out["alias_flags"] == []
    assert out["meta"] == {}
    flags = out["review_flags"]
    assert isinstance(flags, list)
    assert any("hypotheses" in f for f in flags)         # non-list category flagged
    assert any("raw string" in f for f in flags)         # non-dict item flagged
    assert any("init-001" in f for f in flags)           # unhashable id flagged, not crashed


def test_run_ingest_deep_survives_wrong_shape_llm_output(capsys):
    from unittest.mock import patch
    from pmao.deep import run_ingest_deep
    with tempfile.TemporaryDirectory() as tmp:
        vault = _deep_vault(tmp)
        source = vault / "transcripts" / "sync.txt"
        source.write_text("text")
        hostile = {
            "facts": None,
            "action_items": [None, "stray"],
            "initiative_id": ["init-001"],
            "alias_flags": [{"variants": None, "resolved_to": "Dev Patel"}, "stray alias"],
            "review_flags": None,
            "decisions": [{"initiative_id": "init-999", "decision": "orphan"}],
        }
        with patch("pmao.llm.call_structured", return_value=hostile):
            run_ingest_deep(vault, source)    # must not crash on prints or flag appends
        staged_files = list((vault / "staging").glob("*.json"))
        assert len(staged_files) == 1
        out = capsys.readouterr().out
        assert "Dev Patel" in out             # null variants printed without crashing
        assert "init-999" in out              # unknown id still flagged


def test_reconciliation_block_tolerates_hand_edited_files():
    from pathlib import Path
    from pmao.deep import _reconciliation_block
    from pmao.vault import init_vault
    with tempfile.TemporaryDirectory() as tmp:
        vault = Path(tmp)
        init_vault(vault)
        # open action missing 'id' and 'description'; decision missing 'id'
        (vault / "actions.json").write_text(json.dumps([
            {"status": "open", "owner": "Dev Patel"},
        ]))
        (vault / "decisions.json").write_text(json.dumps([
            {"decision": "ship it"},
        ]))
        block = _reconciliation_block(vault)  # no KeyError
        assert "ship it" in block

        # top-level JSON object instead of array — no AttributeError
        (vault / "actions.json").write_text(json.dumps({"act-1": {"status": "open"}}))
        (vault / "decisions.json").write_text(json.dumps({"dec-1": {"decision": "x"}}))
        assert _reconciliation_block(vault) == "none"
