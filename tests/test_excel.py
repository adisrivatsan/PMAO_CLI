import tempfile
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from pmao.models import Initiative
from pmao.excel import create_workbook


def _make_initiative(id_="init-001", **kwargs):
    defaults = dict(
        name="Customer Data Platform",
        status="not_started",
        created=date(2026, 6, 8),
        last_touched=date(2026, 6, 8),
    )
    defaults.update(kwargs)
    return Initiative(id=id_, **defaults)


def test_workbook_has_six_tabs():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        create_workbook(path, [])
        wb = load_workbook(path)
        assert wb.sheetnames == [
            "Initiatives",
            "Full Next Steps Log",
            "Last Touch Log",
            "Meeting Requests",
            "Milestones",
            "Hypotheses",
            "Facts",
            "Signals",
            "Meetings",
            "Review Queue",
        ]


def test_hypotheses_tab_has_headers():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        create_workbook(path, [])
        wb = load_workbook(path)
        ws = wb["Hypotheses"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert headers == ["id", "level", "initiative", "hypothesis", "owner", "validation_path", "status", "created", "source"]


def test_hypotheses_tab_project_level_rows_first():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        init = _make_initiative()
        hyps = [
            {
                "id": "hyp-2026-06-09-0001",
                "initiative_id": "init-001",
                "hypothesis": "Initiative-level hypothesis",
                "owner": "",
                "validation_path": "",
                "status": "open",
                "created": "2026-06-09",
                "source": "call.vtt",
            },
            {
                "id": "hyp-2026-06-09-0000",
                "initiative_id": None,
                "hypothesis": "Project-level hypothesis",
                "owner": "",
                "validation_path": "",
                "status": "open",
                "created": "2026-06-09",
                "source": "call.vtt",
            },
        ]
        create_workbook(path, [init], hypotheses=hyps)
        wb = load_workbook(path)
        ws = wb["Hypotheses"]
        # Row 2 should be the project-level hypothesis (initiative_id None)
        level_col = 2  # "level" is column index 2
        assert ws.cell(2, level_col).value == "Project"
        assert ws.cell(3, level_col).value == "Initiative"


def test_hypotheses_tab_populates_level_column():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        init = _make_initiative()
        hyps = [
            {
                "id": "hyp-2026-06-09-0001",
                "initiative_id": "init-001",
                "hypothesis": "Small practices respond better to outcome-based pricing",
                "owner": "Alex Chen",
                "validation_path": "Run A/B test",
                "status": "open",
                "created": "2026-06-09",
                "source": "strategy-call.vtt",
            },
        ]
        create_workbook(path, [init], hypotheses=hyps)
        wb = load_workbook(path)
        ws = wb["Hypotheses"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(2, headers["level"]).value == "Initiative"
        assert ws.cell(2, headers["initiative"]).value == "Customer Data Platform"
        assert ws.cell(2, headers["hypothesis"]).value == "Small practices respond better to outcome-based pricing"
        assert ws.cell(2, headers["owner"]).value == "Alex Chen"
        assert ws.cell(2, headers["status"]).value == "open"


def test_initiatives_tab_generic_headers():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        create_workbook(path, [])
        wb = load_workbook(path)
        ws = wb["Initiatives"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "initiative" in headers
        assert "coordination_owner" in headers
        assert "responsible_owner" in headers
        assert "current_state" in headers
        assert "coordination_next_steps" in headers
        assert "outstanding_questions" in headers
        assert "outstanding_meetings" in headers
        assert "last_touch_comment" in headers
        assert "last_touch_timestamp" in headers
        # No LRP-specific fields
        assert "elt_presenters" not in headers
        assert "syndication_date" not in headers
        assert "agenda_for_strategy_day" not in headers
        assert "may_28_deadline" not in headers


def test_initiatives_tab_populates_from_initiative():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        init = _make_initiative(
            status="in_progress",
            priority="high",
            coordination_owner="Alex Jordan",
            current_state="Analysis complete",
        )
        create_workbook(path, [init])
        wb = load_workbook(path)
        ws = wb["Initiatives"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        assert ws.cell(2, headers["initiative"]).value == "Customer Data Platform"
        assert ws.cell(2, headers["status"]).value == "in_progress"
        assert ws.cell(2, headers["coordination_owner"]).value == "Alex Jordan"
        assert ws.cell(2, headers["current_state"]).value == "Analysis complete"


def test_full_next_steps_log_headers():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        create_workbook(path, [_make_initiative()])
        wb = load_workbook(path)
        ws = wb["Full Next Steps Log"]
        headers = [ws.cell(1, c).value for c in range(1, 5)]
        assert headers == ["initiative_id", "initiative", "coordination_next_steps", "date_captured"]


def test_full_next_steps_log_populates():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        init = _make_initiative(coordination_next_steps="- Alex to schedule review")
        create_workbook(path, [init])
        wb = load_workbook(path)
        ws = wb["Full Next Steps Log"]
        assert ws.cell(2, 3).value == "- Alex to schedule review"
        assert ws.cell(2, 4).value == "2026-06-08"


def test_last_touch_log_headers():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        create_workbook(path, [_make_initiative()])
        wb = load_workbook(path)
        ws = wb["Last Touch Log"]
        headers = [ws.cell(1, c).value for c in range(1, 5)]
        assert headers == ["initiative_id", "initiative", "last_touch_timestamp", "last_touch_comment"]


def test_last_touch_log_populates():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        init = _make_initiative(
            last_touch_timestamp=date(2026, 6, 7),
            last_touch_comment="[email] Alex confirmed timeline",
        )
        create_workbook(path, [init])
        wb = load_workbook(path)
        ws = wb["Last Touch Log"]
        assert ws.cell(2, 3).value == "2026-06-07"
        assert ws.cell(2, 4).value == "[email] Alex confirmed timeline"


def test_meeting_requests_tab_has_headers():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        create_workbook(path, [])
        wb = load_workbook(path)
        ws = wb["Meeting Requests"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "initiative_ids" in headers
        assert "meeting_title" in headers
        assert "status" in headers


def test_milestones_tab_has_headers():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        create_workbook(path, [])
        wb = load_workbook(path)
        ws = wb["Milestones"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "date" in headers
        assert "title" in headers
        assert "status" in headers


def test_workbook_creates_with_empty_initiatives():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        create_workbook(path, [])
        wb = load_workbook(path)
        ws = wb["Initiatives"]
        # header row only
        assert ws.max_row == 1


def test_workbook_has_deep_extraction_tabs_when_empty():
    import tempfile
    from pathlib import Path
    from openpyxl import load_workbook
    from pmao.excel import create_workbook
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "wb.xlsx"
        create_workbook(path, [])
        wb = load_workbook(path)
        for tab in ("Facts", "Signals", "Meetings", "Review Queue"):
            assert tab in wb.sheetnames, f"missing tab {tab}"
        # Header row present even when empty
        assert wb["Facts"].cell(row=1, column=2).value == "initiative"
        assert wb["Review Queue"].cell(row=1, column=1).value == "staging_file"


def test_workbook_populates_facts_signals_and_review_queue():
    import tempfile
    from datetime import date
    from pathlib import Path
    from openpyxl import load_workbook
    from pmao.excel import create_workbook
    from pmao.models import Initiative

    init = Initiative(id="init-001", name="Pricing", status="in_progress",
                      created=date(2026, 6, 1), last_touched=date(2026, 6, 10))
    facts = [{"id": "fact-0001", "initiative_id": "init-001",
              "claim": "Q2 cost-to-serve up 9%", "stated_by": "Sarah Klein",
              "authority": "owner", "confidence": "high", "inferred": False,
              "source_span": "line 12", "source": "sync.vtt", "created": "2026-06-10"}]
    signals = [{"id": "sig-0001", "initiative_id": "init-001",
                "principal": "Sarah Klein", "lever": "pricing",
                "signal": "wants floor pricing", "implication": "raise floor",
                "source": "sync.vtt", "created": "2026-06-10"}]
    staged = [{"staging_file": "2026-06-10-sync.json", "category": "action_items",
               "initiative_id": "init-001", "summary": "build the cost model"}]
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "wb.xlsx"
        create_workbook(path, [init], facts=facts, signals=signals, staged=staged)
        wb = load_workbook(path)
        assert wb["Facts"].cell(row=2, column=3).value == "Q2 cost-to-serve up 9%"
        assert wb["Facts"].cell(row=2, column=2).value == "Pricing"  # id resolved to name
        assert wb["Signals"].cell(row=2, column=5).value == "wants floor pricing"
        assert wb["Review Queue"].cell(row=2, column=2).value == "action_items"
        assert wb["Review Queue"].cell(row=2, column=4).value == "build the cost model"


def test_meetings_tab_combines_structured_and_initiative_field():
    import tempfile
    from datetime import date
    from pathlib import Path
    from openpyxl import load_workbook
    from pmao.excel import create_workbook
    from pmao.models import Initiative

    init = Initiative(id="init-001", name="Pricing", status="in_progress",
                      created=date(2026, 6, 1), last_touched=date(2026, 6, 10))
    init.outstanding_meetings = "- Kickoff with data team"
    meetings = [{"id": "mtg-0001", "initiative_id": "init-001",
                 "purpose": "align on cost model", "convener": "Sarah Klein",
                 "attendees": ["Finance", "CS"], "functions": ["finance"],
                 "target_timing": "next week", "status": "open",
                 "source": "sync.vtt", "created": "2026-06-10"}]
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "wb.xlsx"
        create_workbook(path, [init], meetings=meetings)
        wb = load_workbook(path)
        ws = wb["Meetings"]
        rows = [[ws.cell(row=r, column=c).value for c in range(1, 11)] for r in (2, 3)]
        purposes = {row[2] for row in rows}
        sources = {row[8] for row in rows}
        assert "align on cost model" in purposes
        assert "- Kickoff with data team" in purposes
        assert sources == {"meetings.json", "initiative-field"}
